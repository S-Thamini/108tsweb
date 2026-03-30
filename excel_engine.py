import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
import shutil

class ExcelEngine:
    def __init__(self):
        pass

    def fast_find(self, ws, labels, exact=False):
        for row in ws.iter_rows(min_row=1, max_row=150, min_col=1, max_col=50):
            for cell in row:
                if cell.value:
                    v = str(cell.value).strip().lower()
                    if exact:
                        if v in [l.lower() for l in labels]: return cell.row, cell.column
                    else:
                        if any(l.lower() in v for l in labels): return cell.row, cell.column
        return None, None

    def run_sync(self, config, data_result, out_dir):
        # 1. Load data
        f = os.path.abspath(config['tpl_file']) # the template file Nifty50.xlsx
        target_sh = config['target_sh']
        idx = config['index']
        exp = config['expiry']
        mid = float(config['mid_strike'])
        rows = int(config['rows'])
        gap = int(config['gap'])
        
        df_fo = data_result.get('fo')
        df_idx = data_result.get('index')
        
        c_sym = next((x for x in ['TckrSymb', 'SYMBOL', 'Symbol'] if x in df_fo.columns), 'SYMBOL')
        c_exp = next((x for x in ['XpryDt', 'EXPIRY_DT', 'Expiry_Dt'] if x in df_fo.columns), 'EXPIRY_DT')
        df_fo[c_sym] = df_fo[c_sym].astype(str).str.strip().str.upper()
        df_fo[c_exp] = df_fo[c_exp].astype(str).str.strip()
        
        idx_df = df_fo[df_fo[c_sym] == idx.upper()]
        if idx_df.empty: raise Exception(f"Symbol '{idx}' not found. Please check data date.")
        nd = idx_df[idx_df[c_exp].str.lower() == exp.lower()]
        if nd.empty: raise Exception(f"Expiry '{exp}' not found in Bhavcopy.")

        # Prepare Map
        c_tp = next((x for x in ['OptnTp', 'OPTION_TYP'] if x in nd.columns), 'OptnTp')
        c_stk = next((x for x in ['StrkPric', 'STRIKE_PR', 'Srike'] if x in nd.columns), 'StrkPric')
        c_o = next((x for x in ['OpnPric', 'OPEN'] if x in nd.columns), 'OpnPric')
        c_h = next((x for x in ['HghPric', 'HIGH'] if x in nd.columns), 'HghPric')
        c_l = next((x for x in ['LwPric', 'LOW'] if x in nd.columns), 'LwPric')
        c_c = next((x for x in ['ClsPric', 'CLOSE'] if x in nd.columns), 'ClsPric')
        df_map = nd.rename(columns={c_tp: 'OptnTp', c_stk: 'StrkPric', c_o: 'OpnPric', c_h: 'HghPric', c_l: 'LwPric', c_c: 'ClsPric'})

        # Index Data
        i_lbl = "NIFTY 50" if idx == "NIFTY" else ("NIFTY BANK" if idx == "BANKNIFTY" else "NIFTY FIN SERVICE")
        idx_row = None
        if df_idx is not None:
            matches = df_idx[df_idx['Index Name'].str.upper() == i_lbl]
            if not matches.empty: idx_row = matches.iloc[0]

        idx_price = 0
        if idx_row is not None:
            c_idx_c = next((x for x in ['Closing Index Value', 'Close'] if x in idx_row.index), None)
            if c_idx_c: idx_price = float(idx_row[c_idx_c])
        if idx_price == 0:
            try: idx_price = float(nd[c_stk].median())
            except: pass

        # 2. Open Workbook using openpyxl
        wb = openpyxl.load_workbook(f)
        
        # Step 1: Use Master Tab Directly (preserves 100% of complex formulas)
        master_sh = config['template']
        if master_sh not in wb.sheetnames:
            master_sh = wb.sheetnames[0] # Fallback to first sheet
            
        ws = wb[master_sh]
        ws.title = target_sh

        # Add initial strikes
        a_r, a_c = self.fast_find(ws, ["strike", "symbol", "srike"])
        if not a_r: a_r, a_c = self.fast_find(ws, ["active contracts"])
        
        if a_r and a_c:
            dr = a_r + 2
            strike_list = [mid + (i * gap) for i in range(-rows, rows + 1)]
            for i, st in enumerate(strike_list):
                ws.cell(row=dr+i, column=a_c).value = st

        # Step 2: Index OHLC Update
        ir, ic = self.fast_find(ws, [i_lbl.lower(), "index", "market update"])
        if not ir and idx.upper() in ["NIFTY", "BANKNIFTY"]:
            ir, ic = 79, 1
        
        if ir and idx_row is not None:
            c_h_idx = 'High Index Value' if 'High Index Value' in idx_row.index else 'High'
            c_l_idx = 'Low Index Value' if 'Low Index Value' in idx_row.index else 'Low'
            c_c_idx = 'Closing Index Value' if 'Closing Index Value' in idx_row.index else 'Close'
            c_o_idx = 'Open Index Value' if 'Open Index Value' in idx_row.index else ('Open' if 'Open' in idx_row.index else 'P.D.Open')
            
            hr = ir - 1
            bold = Font(bold=True)
            align = Alignment(horizontal='center')
            for col_off, name in [(1, 'High'), (2, 'Low'), (3, 'Close'), (5, 'Open')]:
                cell = ws.cell(row=hr, column=ic+col_off)
                cell.value = name
                cell.font = bold
                cell.alignment = align
            
            ws.cell(row=ir, column=ic+1).value = round(float(idx_row[c_h_idx]), 2)
            ws.cell(row=ir, column=ic+2).value = round(float(idx_row[c_l_idx]), 2)
            ws.cell(row=ir, column=ic+3).value = round(float(idx_row[c_c_idx]), 2)
            ws.cell(row=ir, column=ic+5).value = round(float(idx_row[c_o_idx]), 2)

            # Boundaries
            if idx.upper() in ["NIFTY", "BANKNIFTY", "FINNIFTY"]:
                spot_close = float(idx_row[c_c_idx])
                atm_100 = round(spot_close / 100) * 100
                stk_r = ir # simplified fallback
                g_r, g_c = self.fast_find(ws, ["gap down opening"])
                if g_r: stk_r = g_r + 1
                else: 
                    s_r, s_c = self.fast_find(ws, ["srike", "sriike"])
                    if s_r: stk_r = s_r
                
                ws.cell(row=stk_r + 1, column=1).value = atm_100 - 200
                ws.cell(row=stk_r + 2, column=1).value = atm_100 + 200

        # Run Grid processing
        a_r_g, a_c_g = self.fast_find(ws, ["strike", "symbol", "srike"])
        if a_r_g and a_c_g:
            dr_g = a_r_g + 2 if a_r_g == 1 else a_r_g + 1
            for row_idx in range(dr_g, dr_g + 100):
                s_val = ws.cell(row=row_idx, column=a_c_g).value
                if s_val is None or str(s_val).strip() == "": break
                try: s_val = float(s_val)
                except: continue

                c_o_c, c_h_c, c_l_c, c_c_c = a_c_g + 1, a_c_g + 2, a_c_g + 3, a_c_g + 4
                p_o_c, p_h_c, p_l_c, p_c_c = a_c_g + 5, a_c_g + 6, a_c_g + 7, a_c_g + 8

                ce_df = df_map[(abs(df_map['StrkPric'] - s_val) < 0.1) & (df_map['OptnTp'] == 'CE')]
                if not ce_df.empty:
                    ws.cell(row=row_idx, column=c_o_c).value = round(float(ce_df.iloc[0]['OpnPric']), 2)
                    ws.cell(row=row_idx, column=c_h_c).value = round(float(ce_df.iloc[0]['HghPric']), 2)
                    ws.cell(row=row_idx, column=c_l_c).value = round(float(ce_df.iloc[0]['LwPric']), 2)
                    ws.cell(row=row_idx, column=c_c_c).value = round(float(ce_df.iloc[0]['ClsPric']), 2)

                pe_df = df_map[(abs(df_map['StrkPric'] - s_val) < 0.1) & (df_map['OptnTp'] == 'PE')]
                if not pe_df.empty:
                    ws.cell(row=row_idx, column=p_o_c).value = round(float(pe_df.iloc[0]['OpnPric']), 2)
                    ws.cell(row=row_idx, column=p_h_c).value = round(float(pe_df.iloc[0]['HghPric']), 2)
                    ws.cell(row=row_idx, column=p_l_c).value = round(float(pe_df.iloc[0]['LwPric']), 2)
                    ws.cell(row=row_idx, column=p_c_c).value = round(float(pe_df.iloc[0]['ClsPric']), 2)

        # Active Contracts Processing
        a_r, a_c = None, None
        for r_idx in range(70, 150):
            for c_idx in range(1, 20):
                val = str(ws.cell(row=r_idx, column=c_idx).value or "").lower()
                if "active contracts" in val:
                    a_r, a_c = r_idx, c_idx; break
            if a_r: break
            
        if (not a_r) and idx.upper() in ["NIFTY", "BANKNIFTY", "FINNIFTY"]:
            a_r, a_c = 78, 8

        master_active_strikes = []
        if a_r and a_c:
            ws.cell(row=a_r, column=a_c).value = "ACTIVE CONTRACTS"
            ws.cell(row=a_r + 1, column=a_c).value = "option"
            ws.cell(row=a_r + 1, column=a_c + 1).value = "strike"
            ws.cell(row=a_r + 1, column=a_c + 2).value = "strike /low"
            ws.cell(row=a_r + 1, column=a_c + 3).value = "strike/close"
            
            vol_pats = ['TtlTradgVol', 'TtlNbOfTxsExctd', 'TtlTrfVal', 'NoOfCon']
            c_vol = next((x for x in vol_pats if x in df_map.columns), None)
            if c_vol:
                df_map[c_vol] = pd.to_numeric(df_map[c_vol], errors='coerce').fillna(0)
                top6 = df_map.sort_values(c_vol, ascending=False).head(6)
            else:
                top6 = df_map.head(6)

            master_active_strikes = sorted(set(float(s) for s in top6['StrkPric'].unique() if pd.notna(s)))
            dr = a_r + 2
            
            itm_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            otm_fill = PatternFill(start_color="C6EFCE", fill_type="solid")

            for i, (_, r_data) in enumerate(top6.iterrows()):
                tp, stk = r_data.get('OptnTp', 'CE'), float(r_data.get('StrkPric', 0))
                low_p, cls_p = float(r_data.get('LwPric', 0)), float(r_data.get('ClsPric', 0))
                
                if tp == 'CE':
                    is_itm = stk < idx_price - 5
                    is_otm = stk > idx_price + 5
                    status = "ITM" if is_itm else ("OTM" if is_otm else f"ATM ({int(stk)})")
                    s_l, s_c = stk + low_p, stk + cls_p
                else:
                    is_itm = stk > idx_price + 5
                    is_otm = stk < idx_price - 5
                    status = "ITM" if is_itm else ("OTM" if is_otm else f"ATM ({int(stk)})")
                    s_l, s_c = stk - low_p, stk - cls_p
                
                ws.cell(row=dr+i, column=a_c).value = tp
                ws.cell(row=dr+i, column=a_c+1).value = stk
                ws.cell(row=dr+i, column=a_c+2).value = round(s_l, 2)
                ws.cell(row=dr+i, column=a_c+3).value = round(s_c, 2)
                ws.cell(row=dr+i, column=a_c+4).value = status
                
                fill_color = itm_fill if is_itm else otm_fill
                ws.cell(row=dr+i, column=a_c+2).fill = fill_color
                ws.cell(row=dr+i, column=a_c+3).fill = fill_color
                ws.cell(row=dr+i, column=a_c+4).fill = fill_color

        # Straddle
        std_r, std_c = self.fast_find(ws, ["straddle", "stradle"], exact=True)
        if not std_r: std_r, std_c = self.fast_find(ws, ["straddle", "stradle"])
        
        if std_r and std_c:
            curr_r = std_r + 1
            for r_clr in range(std_r + 1, std_r + 22):
                ws.cell(row=r_clr, column=std_c).value = ""
                ws.cell(row=r_clr, column=std_c + 1).value = ""
                ws.cell(row=r_clr, column=std_c + 2).value = ""
                
            for strike in master_active_strikes:
                ce_data = nd[(abs(pd.to_numeric(nd[c_stk], errors='coerce') - strike) < 0.1) & (nd[c_tp] == 'CE')]
                pe_data = nd[(abs(pd.to_numeric(nd[c_stk], errors='coerce') - strike) < 0.1) & (nd[c_tp] == 'PE')]
                if not ce_data.empty and not pe_data.empty:
                    ce_close = float(ce_data.iloc[0].get('ClsPric', 0))
                    pe_close = float(pe_data.iloc[0].get('ClsPric', 0))
                    ws.cell(row=curr_r, column=std_c).value = int(strike) if strike.is_integer() else strike
                    ws.cell(row=curr_r, column=std_c + 1).value = round(strike + ce_close, 2)
                    ws.cell(row=curr_r, column=std_c + 2).value = round(strike - pe_close, 2)
                    ws.cell(row=curr_r + 1, column=std_c + 1).value = round(strike + ce_close + pe_close, 2)
                    ws.cell(row=curr_r + 1, column=std_c + 2).value = round(strike - pe_close - ce_close, 2)
                    curr_r += 3
                    if curr_r > std_r + 20: break

        # ATM Analysis
        atm_r, atm_c = 0, 0
        for r_m in range(1, 150):
            val = str(ws.cell(row=r_m, column=8).value).upper()
            if "ATM ANALYSIS" in val:
                atm_r, atm_c = r_m, 8; break
        if not atm_r: atm_r, atm_c = 90, 8
        
        if atm_r and atm_c:
            stk_vals = pd.to_numeric(nd[c_stk], errors='coerce').dropna()
            if not stk_vals.empty:
                target_spot = idx_price if idx_price > 0 else float(stk_vals.median())
                atm_stk = min(stk_vals.unique(), key=lambda x:abs(x-target_spot))
                for i, t in enumerate(['CE', 'PE']):
                    r_data = nd[(abs(pd.to_numeric(nd[c_stk], errors='coerce') - float(atm_stk)) < 0.1) & (nd[c_tp] == t)]
                    if not r_data.empty:
                        row_v = r_data.iloc[0]
                        ws.cell(row=atm_r+1, column=atm_c+1+i).value = round(float(row_v.get('HghPric', 0)), 2)
                        ws.cell(row=atm_r+2, column=atm_c+1+i).value = round(float(row_v.get('LwPric', 0)), 2)
                        ws.cell(row=atm_r+3, column=atm_c+1+i).value = round(float(row_v.get('ClsPric', 0)), 2)
                        ws.cell(row=atm_r+4, column=atm_c+1+i).value = round(float(row_v.get('OpnPric', 0)), 2)
                t_cell = ws.cell(row=atm_r, column=atm_c)
                if not t_cell.value or "ATM" in str(t_cell.value):
                    t_cell.value = f"ATM ({int(atm_stk)}) Analysis"

        # S/R Max OI
        c_oi = next((x for x in ['OpnIntrst', 'OPEN_INT'] if x in nd.columns), 'OpnIntrst')
        res_r, res_c = self.fast_find(ws, ["res"])
        if res_r: 
            try: ws.cell(row=res_r, column=res_c+1).value = nd[nd[c_tp]=='CE'].sort_values(c_oi).iloc[-1][c_stk]
            except: pass
        sup_r, sup_c = self.fast_find(ws, ["sup"])
        if sup_r: 
            try: ws.cell(row=sup_r, column=sup_c+1).value = nd[nd[c_tp]=='PE'].sort_values(c_oi).iloc[-1][c_stk]
            except: pass

        t_safe = target_sh.replace(" ", "_")[:10]
        out_name = f"108TS_Output_{idx}_{t_safe}.xlsx"
        out_path = os.path.join(out_dir, out_name)
        wb.save(out_path)
        return out_name
